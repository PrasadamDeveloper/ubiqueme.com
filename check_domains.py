#!/usr/bin/env python3
"""
check_domains.py — Lee Dominios-May-15-2026.xlsx, prueba https con y sin www,
y reporta cuáles NO funcionan (cualquier error de carga).
"""

import os
import sys
import time
import ssl
import socket
from urllib.parse import urlparse

import openpyxl
import requests
from requests.exceptions import (
    RequestException,
    Timeout,
    ConnectionError,
    SSLError,
)

# ── Config ──
EXCEL_PATH = os.path.expanduser(
    "/Users/chemex/Documents/domains/Dominios-May-15-2026.xlsx"
)
OUTPUT_PATH = os.path.expanduser(
    "/Users/chemex/Documents/domains/resultados-dominios.xlsx"
)
TIMEOUT = 15  # segundos por petición
MAX_WORKERS = 20  # hilos concurrentes

# Encabezados realistas para evitar bloqueos
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_0) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9,es;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


def extract_domains(excel_path: str) -> list[str]:
    """Extrae dominios únicos de columna B en todas las hojas."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    domains = []
    seen = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
            val = row[0]
            if val is None:
                continue
            raw = str(val).strip().lower()
            # Limpiar: quitar http:// o https:// si alguien los puso
            raw = raw.removeprefix("https://").removeprefix("http://")
            raw = raw.removeprefix("www.")
            raw = raw.split("/")[0]  # quitar rutas
            if raw and raw not in seen:
                seen.add(raw)
                domains.append(raw)
    return domains


def check_url(url: str) -> dict:
    """
    Prueba una URL y devuelve:
      - url
      - ok: True si carga correctamente (status 2xx/3xx con contenido)
      - status_code: int o None
      - error: str resumen del error o None
    """
    result = {"url": url, "ok": False, "status_code": None, "error": None}
    try:
        resp = requests.get(
            url,
            headers=HEADERS,
            timeout=TIMEOUT,
            allow_redirects=True,
            verify=True,
        )
        result["status_code"] = resp.status_code

        # Éxito: 2xx o 3xx con contenido HTML
        if resp.status_code < 400:
            content = resp.text.lower()
            # Detecta páginas de placeholder de hosting (Godaddy, etc.)
            hosting_error_keywords = [
                "sorry! if you are the owner of this website",
                "please contact your hosting provider",
                "there has been a server misconfiguration",
                "the ip address has changed",
                "this domain may have changed recently",
                "the site may have moved to a different server",
                "this site is temporarily unavailable",
                "account has been suspended",
                "parked domain",
                "domain is parked",
                "default web site page",
                "this domain is not configured",
                "website is under construction",
                "welcome to your website",
                "futurosite",
                "futura hosting",
                "pending dns",
                "no hosting",
                "hosting not found",
                "site not found",
                "server not found",
                "no se encuentra el sitio",
                "sitio no configurado",
                "cuenta suspendida",
                "dominio estacionado",
                "en construcción",
                "bienvenido a su sitio",
                "cpanel default",
                "plesk default",
            ]
            is_placeholder = any(kw in content for kw in hosting_error_keywords)

            if is_placeholder:
                result["ok"] = False
                result["error"] = "Página de placeholder/error de hosting"
            elif len(content) < 50:
                result["ok"] = False
                result["error"] = "Respuesta casi vacía"
            else:
                result["ok"] = True
        else:
            result["error"] = f"HTTP {resp.status_code}"

    except SSLError as e:
        result["error"] = f"SSL Error: {str(e)[:80]}"
    except ConnectionError:
        result["error"] = "Conexión rechazada / No reachable"
    except Timeout:
        result["error"] = "Timeout (>{}s)".format(TIMEOUT)
    except socket.gaierror:
        result["error"] = "DNS lookup failed (socket)"
    except RequestException as e:
        result["error"] = str(e)[:120]
    except Exception as e:
        result["error"] = str(e)[:120]

    return result


def check_domain(domain: str) -> list[dict]:
    """Prueba https://dominio y https://www.dominio."""
    variants = [
        f"https://{domain}",
        f"https://www.{domain}",
    ]
    results = []
    for url in variants:
        res = check_url(url)
        results.append(res)
        # Pequeña pausa para no saturar
        time.sleep(0.1)
    return results


def main():
    print("📂 Extrayendo dominios del Excel...")
    domains = extract_domains(EXCEL_PATH)
    print(f"   → {len(domains)} dominios únicos encontrados\n")

    # ── Probar dominios ──
    all_rows = []
    failed_domains = []

    for idx, domain in enumerate(domains, 1):
        print(f"[{idx}/{len(domains)}] Probando {domain}...", end=" ", flush=True)
        results = check_domain(domain)

        # Determinar estado consolidado del dominio
        any_ok = any(r["ok"] for r in results)
        status_domain = "✅ OK" if any_ok else "❌ FALLA"

        if not any_ok:
            failed_domains.append(domain)

        for r in results:
            label = "www" if "www." in r["url"] else "sin-www"
            ok_label = "✅" if r["ok"] else "❌"
            print(f"  {label}: {ok_label} {r.get('status_code') or ''} {r.get('error') or ''}")
            all_rows.append(
                {
                    "Dominio": domain,
                    "Variante": label,
                    "URL": r["url"],
                    "Status": ok_label,
                    "Codigo HTTP": r.get("status_code") or "",
                    "Error": r.get("error") or "",
                }
            )

        if any_ok:
            print(f"  → {status_domain}")
        else:
            print(f"  → {status_domain}")
        print()

    # ── Guardar Excel ──
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Resultados"

    headers = ["Dominio", "Variante", "URL", "Status", "Codigo HTTP", "Error"]
    ws_out.append(headers)

    for row in all_rows:
        ws_out.append([row[h] for h in headers])

    # Autoajustar ancho
    for col_idx, header in enumerate(headers, 1):
        ws_out.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(
            12, len(header) + 4
        )

    # ── Hoja resumen ──
    ws_summary = wb_out.create_sheet("Resumen")
    ws_summary.append(["Dominio", "Estado"])
    for d in domains:
        any_ok = any(
            row["Dominio"] == d and row["Status"] == "✅"
            for row in all_rows
        )
        ws_summary.append([d, "✅ OK" if any_ok else "❌ FALLA"])

    wb_out.save(OUTPUT_PATH)
    print(f"\n📁 Resultados guardados en: {OUTPUT_PATH}")

    # ── Reporte en consola ──
    print(f"\n{'='*60}")
    print(f"📊 RESUMEN FINAL")
    print(f"{'='*60}")
    print(f"Total dominios:      {len(domains)}")
    print(f"Funcionando:         {len(domains) - len(failed_domains)}")
    print(f"Fallando:            {len(failed_domains)}")

    if failed_domains:
        print(f"\n{'='*60}")
        print("❌ DOMINIOS QUE FALLARON:")
        print(f"{'='*60}")
        for d in failed_domains:
            print(f"  - {d}")

    return 0 if not failed_domains else 1


if __name__ == "__main__":
    sys.exit(main())